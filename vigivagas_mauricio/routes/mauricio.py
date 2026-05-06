from io import BytesIO
from urllib.parse import urlencode

from flask import Blueprint, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import check_password_hash

from utils.auth import mauricio_required
from utils.db import get_connection
from utils.audit import log_action
from utils.validators import normalize_upper
from utils.privacy import mask_cnpj, mask_email, mask_ip, mask_phone

STATUS_RECRUTADOR = {"pendente", "validado", "verificado"}
STATUS_ANTIFRAUDE = {"normal", "suspeito", "bloqueado"}

mauricio_bp = Blueprint("mauricio", __name__, url_prefix="/mauricio")


def _dashboard_redirect(anchor: str = ""):
    params = request.args.to_dict(flat=True)
    target = url_for("mauricio.dashboard", **params)
    return redirect(f"{target}#{anchor}" if anchor else target)


def _require_mauricio_password() -> bool:
    senha = request.form.get("mauricio_password", "")
    if not senha:
        flash("Informe sua senha do painel para confirmar esta ação sensível.", "error")
        return False
    conn = get_connection()
    usuario = conn.execute("SELECT password FROM mauricio_usuarios WHERE id = ? AND ativo = 1", (session.get("mauricio_id"),)).fetchone()
    conn.close()
    if not usuario or not check_password_hash(usuario["password"], senha):
        flash("Senha do painel inválida. Ação não executada.", "error")
        return False
    return True


def _build_workbook(sheet_name: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill(fill_type="solid", fgColor="0A4F93")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(18, min(42, len(header) + 4))
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value="" if value is None else value)
    ws.freeze_panes = "A2"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _parse_limit(args, key: str, default: int = 10) -> int:
    try:
        value = int(args.get(key, default))
    except (TypeError, ValueError):
        value = default
    return max(5, min(value, 200))


def _build_next_link(args, key: str, step: int, anchor: str) -> str:
    params = args.to_dict(flat=True)
    params[key] = str(_parse_limit(args, key) + step)
    return f"{url_for('mauricio.dashboard')}?{urlencode(params)}#{anchor}"


def _fetch_dashboard_data(args):
    filtro_v_status = args.get("v_status", "").strip().lower()
    filtro_v_cidade = normalize_upper(args.get("v_cidade", ""))
    filtro_v_nome = normalize_upper(args.get("v_nome", ""))
    filtro_v_antifraude = args.get("v_antifraude", "").strip().lower()

    filtro_r_status = args.get("r_status", "").strip().lower()
    filtro_r_nome_empresa = normalize_upper(args.get("r_nome_empresa", ""))
    filtro_r_email = args.get("r_email", "").strip().lower()
    filtro_r_antifraude = args.get("r_antifraude", "").strip().lower()

    filtro_g_empresa = normalize_upper(args.get("g_empresa", ""))
    filtro_g_cidade = normalize_upper(args.get("g_cidade", ""))
    filtro_g_status = args.get("g_status", "").strip().lower()

    filtro_c_status = args.get("c_status", "").strip().lower()
    filtro_c_nome = normalize_upper(args.get("c_nome", ""))
    filtro_c_empresa = normalize_upper(args.get("c_empresa", ""))
    filtro_c_cidade = normalize_upper(args.get("c_cidade", ""))

    r_limit = _parse_limit(args, "r_limit")
    v_limit = _parse_limit(args, "v_limit")
    g_limit = _parse_limit(args, "g_limit")
    c_limit = _parse_limit(args, "c_limit")

    conn = get_connection()
    resumo = {
        "recrutadores": conn.execute("SELECT COUNT(*) AS total FROM recrutadores").fetchone()["total"],
        "recrutadores_pendentes": conn.execute("SELECT COUNT(*) AS total FROM recrutadores WHERE LOWER(status) = 'pendente'").fetchone()["total"],
        "recrutadores_validados": conn.execute("SELECT COUNT(*) AS total FROM recrutadores WHERE LOWER(status) IN ('validado', 'ativo')").fetchone()["total"],
        "recrutadores_verificados": conn.execute("SELECT COUNT(*) AS total FROM recrutadores WHERE LOWER(status) = 'verificado'").fetchone()["total"],
        "vigilantes": conn.execute("SELECT COUNT(*) AS total FROM vigilantes").fetchone()["total"],
        "candidaturas": conn.execute("SELECT COUNT(*) AS total FROM candidaturas").fetchone()["total"],
        "vagas": conn.execute("SELECT COUNT(*) AS total FROM vagas").fetchone()["total"],
        "cadastros_suspeitos": conn.execute("SELECT COALESCE((SELECT COUNT(*) FROM vigilantes WHERE LOWER(antifraude_status) = 'suspeito'), 0) + COALESCE((SELECT COUNT(*) FROM recrutadores WHERE LOWER(antifraude_status) = 'suspeito'), 0) AS total").fetchone()["total"],
        "cadastros_bloqueados": conn.execute("SELECT COALESCE((SELECT COUNT(*) FROM vigilantes WHERE LOWER(antifraude_status) = 'bloqueado'), 0) + COALESCE((SELECT COUNT(*) FROM recrutadores WHERE LOWER(antifraude_status) = 'bloqueado'), 0) AS total").fetchone()["total"],
    }

    where_r, params_r = [], []
    if filtro_r_status:
        if filtro_r_status == "validado":
            where_r.append("LOWER(status) IN ('validado', 'ativo')")
        else:
            where_r.append("LOWER(status) = ?")
            params_r.append(filtro_r_status)
    if filtro_r_nome_empresa:
        where_r.append("(nome_empresa LIKE ? OR nome_responsavel LIKE ?)")
        params_r.extend([f"%{filtro_r_nome_empresa}%", f"%{filtro_r_nome_empresa}%"])
    if filtro_r_email:
        where_r.append("LOWER(email) LIKE ?")
        params_r.append(f"%{filtro_r_email}%")
    if filtro_r_antifraude:
        where_r.append("LOWER(antifraude_status) = ?")
        params_r.append(filtro_r_antifraude)
    recrutadores_query = """
        SELECT id, nome_responsavel, nome_empresa, email, telefone, cidade, estado, cnpj, razao_social,
               situacao_cadastral, site_empresa, descricao_empresa, cnpj_modo_validacao,
               email_verificado, status, created_at, antifraude_score, antifraude_flags,
               antifraude_status, ip_cadastro
        FROM recrutadores
    """
    if where_r:
        recrutadores_query += " WHERE " + " AND ".join(where_r)
    recrutadores_query += " ORDER BY CASE LOWER(status) WHEN 'pendente' THEN 0 WHEN 'validado' THEN 1 WHEN 'ativo' THEN 1 WHEN 'verificado' THEN 2 ELSE 3 END, id DESC LIMIT ?"
    recrutadores = conn.execute(recrutadores_query, [*params_r, r_limit]).fetchall()

    where_v, params_v = [], []
    if filtro_v_status:
        if filtro_v_status == "validado":
            where_v.append("LOWER(status) IN ('validado', 'ativo')")
        else:
            where_v.append("LOWER(status) = ?")
            params_v.append(filtro_v_status)
    if filtro_v_cidade:
        where_v.append("cidade LIKE ?")
        params_v.append(f"%{filtro_v_cidade}%")
    if filtro_v_nome:
        where_v.append("(nome LIKE ? OR email LIKE ? OR telefone LIKE ? OR instituicao_formacao LIKE ?)")
        params_v.extend([f"%{filtro_v_nome}%", f"%{filtro_v_nome.lower()}%", f"%{filtro_v_nome}%", f"%{filtro_v_nome}%"])
    if filtro_v_antifraude:
        where_v.append("LOWER(antifraude_status) = ?")
        params_v.append(filtro_v_antifraude)
    vigilantes_query = "SELECT id, nome, cidade, estado, email, telefone, cep, endereco, escolaridade, possui_cfv, instituicao_formacao, data_ultima_reciclagem, curso_ultima_reciclagem, ultima_experiencia_profissional, status, antifraude_score, antifraude_flags, antifraude_status, ip_cadastro, created_at FROM vigilantes"
    if where_v:
        vigilantes_query += " WHERE " + " AND ".join(where_v)
    vigilantes_query += " ORDER BY CASE LOWER(antifraude_status) WHEN 'bloqueado' THEN 0 WHEN 'suspeito' THEN 1 ELSE 2 END, id DESC LIMIT ?"
    vigilantes = conn.execute(vigilantes_query, [*params_v, v_limit]).fetchall()

    where_g, params_g = [], []
    if filtro_g_empresa:
        where_g.append("(v.empresa LIKE ? OR COALESCE(r.nome_empresa,'') LIKE ?)")
        params_g.extend([f"%{filtro_g_empresa}%", f"%{filtro_g_empresa}%"])
    if filtro_g_cidade:
        where_g.append("v.cidade LIKE ?")
        params_g.append(f"%{filtro_g_cidade}%")
    if filtro_g_status:
        where_g.append("LOWER(v.status) = ?")
        params_g.append(filtro_g_status)
    vagas_query = """
        SELECT v.id, v.titulo, v.empresa, v.cidade, v.estado, v.status, v.created_at,
               r.nome_empresa AS recrutador_empresa,
               COUNT(c.id) AS total_candidaturas
        FROM vagas v
        LEFT JOIN recrutadores r ON r.id = v.recrutador_id
        LEFT JOIN candidaturas c ON c.vaga_id = v.id
    """
    if where_g:
        vagas_query += " WHERE " + " AND ".join(where_g)
    vagas_query += " GROUP BY v.id, v.titulo, v.empresa, v.cidade, v.estado, v.status, v.created_at, r.nome_empresa ORDER BY v.id DESC LIMIT ?"
    vagas = conn.execute(vagas_query, [*params_g, g_limit]).fetchall()

    where_c, params_c = [], []
    if filtro_c_status:
        where_c.append("LOWER(c.status) = ?")
        params_c.append(filtro_c_status)
    if filtro_c_nome:
        where_c.append("(vig.nome LIKE ? OR vig.email LIKE ?)")
        params_c.extend([f"%{filtro_c_nome}%", f"%{filtro_c_nome.lower()}%"])
    if filtro_c_empresa:
        where_c.append("(v.empresa LIKE ? OR COALESCE(r.nome_empresa,'') LIKE ?)")
        params_c.extend([f"%{filtro_c_empresa}%", f"%{filtro_c_empresa}%"])
    if filtro_c_cidade:
        where_c.append("(vig.cidade LIKE ? OR v.cidade LIKE ?)")
        params_c.extend([f"%{filtro_c_cidade}%", f"%{filtro_c_cidade}%"])
    candidaturas_query = """
        SELECT c.id, c.status, c.observacoes, c.created_at, c.updated_at,
               vig.nome AS vigilante_nome, vig.email AS vigilante_email, vig.telefone AS vigilante_telefone, vig.cidade AS vigilante_cidade, vig.estado AS vigilante_estado,
               v.id AS vaga_id, v.titulo AS vaga_titulo, v.empresa AS vaga_empresa, v.cidade AS vaga_cidade, v.estado AS vaga_estado,
               COALESCE(r.nome_empresa, v.empresa) AS empresa_origem
        FROM candidaturas c
        JOIN vigilantes vig ON vig.id = c.vigilante_id
        JOIN vagas v ON v.id = c.vaga_id
        LEFT JOIN recrutadores r ON r.id = v.recrutador_id
    """
    if where_c:
        candidaturas_query += " WHERE " + " AND ".join(where_c)
    candidaturas_query += " ORDER BY c.id DESC LIMIT ?"
    candidaturas = conn.execute(candidaturas_query, [*params_c, c_limit]).fetchall()

    oportunidades_encerradas = conn.execute("""
        SELECT v.id, v.titulo, v.empresa, v.cidade, v.estado, v.status, v.encerrada_em,
               v.encerrada_motivo_tipo, v.encerrada_motivo, v.vigivagas_ajudou_contratacao,
               v.contratacoes_quantidade, v.encerrada_observacoes, v.created_at,
               COALESCE(r.nome_empresa, v.empresa) AS recrutador_empresa,
               r.nome_responsavel, r.email AS recrutador_email,
               COUNT(c.id) AS total_candidaturas
        FROM vagas v
        LEFT JOIN recrutadores r ON r.id = v.recrutador_id
        LEFT JOIN candidaturas c ON c.vaga_id = v.id
        WHERE v.status = 'encerrada'
        GROUP BY v.id, v.titulo, v.empresa, v.cidade, v.estado, v.status, v.encerrada_em,
                 v.encerrada_motivo_tipo, v.encerrada_motivo, v.vigivagas_ajudou_contratacao,
                 v.contratacoes_quantidade, v.encerrada_observacoes, v.created_at,
                 r.nome_empresa, r.nome_responsavel, r.email
        ORDER BY CASE WHEN v.encerrada_em IS NULL THEN 1 ELSE 0 END, v.encerrada_em DESC, v.created_at DESC
        LIMIT 50
    """).fetchall()

    audit_logs = conn.execute("""
        SELECT actor_type, actor_id, action, entity_type, entity_id, details, ip_address, created_at
        FROM audit_logs
        ORDER BY id DESC
        LIMIT 20
    """).fetchall()
    lgpd_requests = conn.execute("""
        SELECT id, user_type, user_id, email, request_type, status, details, ip_address, created_at, resolved_at
        FROM lgpd_requests
        ORDER BY id DESC
        LIMIT 20
    """).fetchall()
    conn.close()

    filtros = {
        "v_status": filtro_v_status,
        "v_cidade": filtro_v_cidade,
        "v_nome": filtro_v_nome,
        "v_antifraude": filtro_v_antifraude,
        "r_status": filtro_r_status,
        "r_nome_empresa": filtro_r_nome_empresa,
        "r_email": filtro_r_email,
        "r_antifraude": filtro_r_antifraude,
        "g_empresa": filtro_g_empresa,
        "g_cidade": filtro_g_cidade,
        "g_status": filtro_g_status,
        "c_status": filtro_c_status,
        "c_nome": filtro_c_nome,
        "c_empresa": filtro_c_empresa,
        "c_cidade": filtro_c_cidade,
    }
    limites = {"r_limit": r_limit, "v_limit": v_limit, "g_limit": g_limit, "c_limit": c_limit}
    return resumo, recrutadores, vigilantes, vagas, candidaturas, filtros, limites, audit_logs, lgpd_requests, oportunidades_encerradas


@mauricio_bp.route("/")
def root():
    return redirect(url_for("mauricio.login"))


@mauricio_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_connection()
        usuario = conn.execute(
            "SELECT * FROM mauricio_usuarios WHERE username = ? AND ativo = 1",
            (username,),
        ).fetchone()
        conn.close()
        if usuario and check_password_hash(usuario["password"], password):
            session["mauricio_id"] = usuario["id"]
            session["mauricio_nome"] = usuario["nome_exibicao"] or usuario["username"]
            return redirect(url_for("mauricio.dashboard"))
        flash("Usuário ou senha inválidos para o painel do Maurício.", "error")
    return render_template("mauricio/login.html")


@mauricio_bp.route("/dashboard")
@mauricio_required
def dashboard():
    log_action("mauricio", session.get("mauricio_id"), "acessou_dashboard", "painel", None, "Acesso ao painel administrativo com dados pessoais mascarados")
    resumo, recrutadores, vigilantes, vagas, candidaturas, filtros, limites, audit_logs, lgpd_requests, oportunidades_encerradas = _fetch_dashboard_data(request.args)
    next_links = {
        "recrutadores": _build_next_link(request.args, "r_limit", 10, "secao-recrutadores"),
        "vigilantes": _build_next_link(request.args, "v_limit", 10, "secao-vigilantes"),
        "vagas": _build_next_link(request.args, "g_limit", 10, "secao-vagas"),
        "candidaturas": _build_next_link(request.args, "c_limit", 10, "secao-candidaturas"),
    }
    return render_template(
        "mauricio/dashboard.html",
        resumo=resumo,
        recrutadores=recrutadores,
        vigilantes=vigilantes,
        vagas=vagas,
        candidaturas=candidaturas,
        filtros=filtros,
        limites=limites,
        next_links=next_links,
        audit_logs=audit_logs,
        lgpd_requests=lgpd_requests,
        oportunidades_encerradas=oportunidades_encerradas,
        mask_email=mask_email,
        mask_phone=mask_phone,
        mask_ip=mask_ip,
        mask_cnpj=mask_cnpj,
    )


@mauricio_bp.route("/exportar/recrutadores.xlsx")
@mauricio_required
def exportar_recrutadores():
    _, recrutadores, _, _, _, _, _, _, _, _ = _fetch_dashboard_data(request.args)
    rows = [[r["id"], r["nome_responsavel"], r["nome_empresa"], mask_email(r["email"]), mask_phone(r["telefone"]), r["cidade"], r["estado"], mask_cnpj(r["cnpj"]), r["razao_social"], r["situacao_cadastral"], r["email_verificado"], r["status"], r["antifraude_status"], r["antifraude_score"], r["antifraude_flags"], r["created_at"]] for r in recrutadores]
    output = _build_workbook("Recrutadores", ["ID", "Responsável", "Empresa", "E-mail mascarado", "Telefone mascarado", "Cidade", "Estado", "CNPJ mascarado", "Razão Social", "Situação Cadastral", "E-mail Verificado", "Status", "Antifraude", "Score", "Flags", "Criado em"], rows)
    log_action("mauricio", session.get("mauricio_id"), "exportou_xlsx", "recrutadores", None, f"Exportação de recrutadores.xlsx com {len(rows)} linhas; e-mail, telefone e CNPJ mascarados")
    return send_file(output, as_attachment=True, download_name="vigivagas_recrutadores.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@mauricio_bp.route("/exportar/vigilantes.xlsx")
@mauricio_required
def exportar_vigilantes():
    _, _, vigilantes, _, _, _, _, _, _, _ = _fetch_dashboard_data(request.args)
    rows = [[r["id"], r["nome"], r["cidade"], r["estado"], "REMOVIDO", "REMOVIDO", mask_email(r["email"]), mask_phone(r["telefone"]), r["escolaridade"], r["possui_cfv"], r["instituicao_formacao"], r["data_ultima_reciclagem"], r["curso_ultima_reciclagem"], r["ultima_experiencia_profissional"], r["status"], r["antifraude_status"], r["antifraude_score"], r["antifraude_flags"], r["created_at"]] for r in vigilantes]
    output = _build_workbook("Vigilantes", ["ID", "Nome", "Cidade", "Estado", "CEP", "Endereço", "E-mail mascarado", "Telefone mascarado", "Escolaridade", "Possui CFV", "Instituição de Formação", "Data Última Reciclagem", "Curso Última Reciclagem", "Última Experiência", "Status", "Antifraude", "Score", "Flags", "Criado em"], rows)
    log_action("mauricio", session.get("mauricio_id"), "exportou_xlsx", "vigilantes", None, f"Exportação de vigilantes.xlsx com {len(rows)} linhas; contato mascarado e endereço removido")
    return send_file(output, as_attachment=True, download_name="vigivagas_vigilantes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@mauricio_bp.route("/exportar/vagas.xlsx")
@mauricio_required
def exportar_vagas():
    _, _, _, vagas, _, _, _, _, _, _ = _fetch_dashboard_data(request.args)
    rows = [[r["id"], r["titulo"], r["empresa"], r["recrutador_empresa"], r["cidade"], r["estado"], r["status"], r["total_candidaturas"], r["created_at"]] for r in vagas]
    output = _build_workbook("Vagas", ["ID", "Título", "Empresa", "Empresa do Recrutador", "Cidade", "Estado", "Status", "Total de Candidaturas", "Criado em"], rows)
    log_action("mauricio", session.get("mauricio_id"), "exportou_xlsx", "vagas", None, f"Exportação de vagas.xlsx com {len(rows)} linhas")
    return send_file(output, as_attachment=True, download_name="vigivagas_vagas.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@mauricio_bp.route("/exportar/candidaturas.xlsx")
@mauricio_required
def exportar_candidaturas():
    _, _, _, _, candidaturas, _, _, _, _, _ = _fetch_dashboard_data(request.args)
    rows = [[r["id"], r["vigilante_nome"], mask_email(r["vigilante_email"]), mask_phone(r["vigilante_telefone"]), r["vigilante_cidade"], r["vigilante_estado"], r["vaga_id"], r["vaga_titulo"], r["vaga_empresa"], r["vaga_cidade"], r["vaga_estado"], r["status"], r["observacoes"], r["created_at"], r["updated_at"]] for r in candidaturas]
    output = _build_workbook("Candidaturas", ["ID", "Vigilante", "E-mail mascarado", "Telefone mascarado", "Cidade Vigilante", "Estado Vigilante", "Vaga ID", "Título da Vaga", "Empresa", "Cidade da Vaga", "Estado da Vaga", "Status", "Observações", "Criado em", "Atualizado em"], rows)
    log_action("mauricio", session.get("mauricio_id"), "exportou_xlsx", "candidaturas", None, f"Exportação de candidaturas.xlsx com {len(rows)} linhas; contatos mascarados e CPF ausente")
    return send_file(output, as_attachment=True, download_name="vigivagas_candidaturas.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@mauricio_bp.route("/exportar/vaga/<int:vaga_id>/candidaturas.xlsx")
@mauricio_required
def exportar_candidaturas_vaga(vaga_id: int):
    conn = get_connection()
    vaga = conn.execute("SELECT id, titulo, empresa FROM vagas WHERE id = ?", (vaga_id,)).fetchone()
    if not vaga:
        conn.close()
        flash("Vaga não encontrada para exportação.", "error")
        return redirect(url_for("mauricio.dashboard"))
    candidaturas = conn.execute(
        """
        SELECT c.id, c.status, c.observacoes, c.created_at, c.updated_at,
               v.nome, v.cpf, v.telefone, v.email, v.cidade, v.curso, v.reciclagem
        FROM candidaturas c
        JOIN vigilantes v ON v.id = c.vigilante_id
        WHERE c.vaga_id = ?
        ORDER BY c.id DESC
        """,
        (vaga_id,),
    ).fetchall()
    conn.close()
    rows = [[r["id"], r["nome"], mask_phone(r["telefone"]), mask_email(r["email"]), r["cidade"], r["curso"], r["reciclagem"], r["status"], r["observacoes"], r["created_at"], r["updated_at"]] for r in candidaturas]
    output = _build_workbook("Candidaturas", ["Candidatura ID", "Nome", "Telefone mascarado", "E-mail mascarado", "Cidade", "Curso", "Reciclagem", "Status", "Observações", "Criado em", "Atualizado em"], rows)
    log_action("mauricio", session.get("mauricio_id"), "exportou_xlsx", "candidaturas_vaga", vaga_id, f"Exportação de candidaturas da vaga {vaga_id} com {len(rows)} linhas; CPF removido da planilha")
    return send_file(output, as_attachment=True, download_name=f"vigivagas_vaga_{vaga_id}_candidaturas.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@mauricio_bp.route("/recrutadores/<int:recrutador_id>/status", methods=["POST"])
@mauricio_required
def atualizar_status_recrutador(recrutador_id: int):
    novo_status = request.form.get("status", "").strip().lower()
    if novo_status not in STATUS_RECRUTADOR:
        flash("Status de recrutador inválido.", "error")
        return _dashboard_redirect("secao-recrutadores")
    if not _require_mauricio_password():
        return _dashboard_redirect("secao-recrutadores")
    conn = get_connection()
    recrutador = conn.execute("SELECT id, nome_empresa, nome_responsavel FROM recrutadores WHERE id = ?", (recrutador_id,)).fetchone()
    if not recrutador:
        conn.close()
        flash("Recrutador não encontrado.", "error")
        return redirect(url_for("mauricio.dashboard"))
    conn.execute("UPDATE recrutadores SET status = ? WHERE id = ?", (novo_status, recrutador_id))
    conn.commit()
    conn.close()
    log_action("mauricio", session.get("mauricio_id"), "alterou_status_recrutador", "recrutador", recrutador_id, f"Novo status: {novo_status}")
    flash(f"Status de {recrutador['nome_responsavel']} / {recrutador['nome_empresa'] or 'SEM EMPRESA'} alterado para {novo_status.upper()}.", "success")
    return _dashboard_redirect("secao-recrutadores")


@mauricio_bp.route("/recrutadores/<int:recrutador_id>/antifraude", methods=["POST"])
@mauricio_required
def atualizar_antifraude_recrutador(recrutador_id: int):
    novo_status = request.form.get("antifraude_status", "").strip().lower()
    if novo_status not in STATUS_ANTIFRAUDE:
        flash("Status antifraude do recrutador inválido.", "error")
        return _dashboard_redirect("secao-recrutadores")
    if not _require_mauricio_password():
        return _dashboard_redirect("secao-recrutadores")
    conn = get_connection()
    recrutador = conn.execute("SELECT id, nome_empresa, nome_responsavel FROM recrutadores WHERE id = ?", (recrutador_id,)).fetchone()
    if not recrutador:
        conn.close()
        flash("Recrutador não encontrado.", "error")
        return redirect(url_for("mauricio.dashboard"))
    conn.execute("UPDATE recrutadores SET antifraude_status = ? WHERE id = ?", (novo_status, recrutador_id))
    conn.commit()
    conn.close()
    log_action("mauricio", session.get("mauricio_id"), "alterou_antifraude_recrutador", "recrutador", recrutador_id, f"Novo antifraude: {novo_status}")
    flash(f"Antifraude de {recrutador['nome_responsavel']} / {recrutador['nome_empresa'] or 'SEM EMPRESA'} alterado para {novo_status.upper()}.", "success")
    return _dashboard_redirect("secao-recrutadores")


@mauricio_bp.route("/vigilantes/<int:vigilante_id>/antifraude", methods=["POST"])
@mauricio_required
def atualizar_antifraude_vigilante(vigilante_id: int):
    novo_status = request.form.get("antifraude_status", "").strip().lower()
    if novo_status not in STATUS_ANTIFRAUDE:
        flash("Status antifraude do vigilante inválido.", "error")
        return _dashboard_redirect("secao-vigilantes")
    if not _require_mauricio_password():
        return _dashboard_redirect("secao-vigilantes")
    conn = get_connection()
    vigilante = conn.execute("SELECT id, nome, email FROM vigilantes WHERE id = ?", (vigilante_id,)).fetchone()
    if not vigilante:
        conn.close()
        flash("Vigilante não encontrado.", "error")
        return redirect(url_for("mauricio.dashboard"))
    conn.execute("UPDATE vigilantes SET antifraude_status = ? WHERE id = ?", (novo_status, vigilante_id))
    conn.commit()
    conn.close()
    log_action("mauricio", session.get("mauricio_id"), "alterou_antifraude_vigilante", "vigilante", vigilante_id, f"Novo antifraude: {novo_status}")
    flash(f"Antifraude de {vigilante['nome']} / {vigilante['email']} alterado para {novo_status.upper()}.", "success")
    return _dashboard_redirect("secao-vigilantes")


@mauricio_bp.route("/logout", methods=["POST"])
@mauricio_required
def logout():
    session.pop("mauricio_id", None)
    session.pop("mauricio_nome", None)
    flash("Sessão do Maurício encerrada com sucesso.", "success")
    return redirect(url_for("mauricio.login"))
